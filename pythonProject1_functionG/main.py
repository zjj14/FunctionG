import FunctiondPdG
import Maxmium_findnew
import fanyan
import extract
import getwidth
import getLie
from openpyxl import load_workbook
import openpyxl

import water_hammer
from FunctionG import FunctionG
from Functionthe import Functionthe
import xiaobochange
import liqun
# 打开 Excel 文件,将文件内容读取到extract_data列表中
file_path = 'C:\\Users\\张骏\\Desktop\Data\\J33.xlsx'

extract_data=extract.extract(file_path)
#下面将每一列数据进行分割
data_width=getwidth.get_list_width(extract_data)
print(data_width)
i=0
sum=0
repairindex=[]
while i<data_width:
    the=[]
    p=[]
    a=[]
    b=[]
    G=[]
    wh=[]
    GdPdG=[]
    a=getLie.getLie(extract_data,i)
    b=getLie.getLie(extract_data,i+1)
    i = i + 2  # 位置问题
    for j in range(len(a)):
        if a[j]!=0:
            the.append(a[j])
            p.append(b[j])
    if len(the)<30:#小于三十行的数据不要了
        continue
    sum=sum+1
    path = f'C:\\Users\\张骏\\Desktop\\Data\\xiaobo_test\\{sum + 1}.xlsx'
    # 创建一个新的 Excel 工作簿
    workbook = openpyxl.Workbook()
    # 选择默认的工作表
    sheet1 = workbook.active
    workbook.save(path)
#取出压力和时间数据，并找出压力数据是否有极大值，找出最后一个极大值并反改前面的数据
    theta=Functionthe(the)
    G=FunctionG(theta)
    result = Maxmium_findnew.Maxmuim(G, p)#看压力的极大值这里应该以G函数为横坐标还是以记录的时间为横坐标
    wh=water_hammer.findwh(p)
    max_the = result[0]
    max_p = result[1]
    max_temp = result[2]
    if max_temp>=len(p)-6:
        continue
    print(max_temp)

    if max_temp != 0:
        updated_p = fanyan.update_values(G, p, max_temp)#####
        repairindex.append(i)
    else:
        updated_p=p
    dPdG=FunctiondPdG.FunctiondPdG(updated_p,G)#通过函数得到dPdG
    print(len(dPdG))
    print(len(G))
    print(len(updated_p))
    for z in range(len(G)):
        GdPdG.append(G[z] * dPdG[z])
    GdPdG_liqun=list(liqun.deleteliqun(GdPdG,0.3))
    GdPdG_change=list(xiaobochange.wavelet_denoising(GdPdG_liqun))#####
    print(type(GdPdG_change))
    print(len(GdPdG_change))
    print(GdPdG_change)
    combined_list = [theta,G, updated_p, dPdG, GdPdG_change,wh]
    #combined_list=zip(combined_list)
    #print(combined_list)
    #writebook= openpyxl.Workbook()
    #sheet1=writebook.active
    for row in combined_list:
        sheet1.append(row)
    workbook.save(path)

#先用main运行读取数据，再运行sorted给文件夹中的excel文件进行编号，再使用images画图

